import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from datetime import datetime
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from matplotlib.patches import Patch
import unicodedata

# Configuración de la página
st.set_page_config(
    page_title="Tablero resumen de gestión de Autorizaciones y Programación en Tramita",
    page_icon="📊",
    layout="wide"
)

# Estilos CSS personalizados para dashboard
st.markdown("""
    <style>
    .metric-card {
        background: linear-gradient(135deg, #7c3aed, #6d28d9);
        padding: 20px;
        border-radius: 10px;
        color: white;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .metric-card .metric-value {
        font-size: 32px !important;
        font-weight: bold;
        margin: 5px 0 0 0;
    }
    .metric-card .metric-label {
        font-size: 14px;
        opacity: 0.9;
        margin: 0;
    }
    .metric-card-small {
        background: linear-gradient(135deg, #8b5cf6, #7c3aed);
        padding: 15px;
        border-radius: 10px;
        color: white;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .metric-card-small .metric-value {
        font-size: 26px !important;
        font-weight: bold;
        margin: 5px 0 0 0;
    }
    .metric-card-small .metric-label {
        font-size: 12px;
        opacity: 0.9;
        margin: 0;
    }
    .chart-container {
        background: white;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        margin-bottom: 20px;
        border: 1px solid #f0f0f0;
    }
    .interpretation-box {
        background: #f8f4ff;
        padding: 15px 20px;
        border-radius: 8px;
        border-left: 4px solid #7c3aed;
        margin-top: 10px;
        color: #2d2d2d;
        font-size: 14px;
        line-height: 1.6;
    }
    .interpretation-box strong {
        color: #5b21b6;
    }
    .executive-summary {
        background: linear-gradient(135deg, #f8f4ff, #ede9fe);
        padding: 25px 30px;
        border-radius: 12px;
        border: 2px solid #7c3aed;
        margin: 20px 0 30px 0;
        color: #2d2d2d;
        font-size: 15px;
        line-height: 1.8;
    }
    .executive-summary h3 {
        color: #5b21b6;
        margin-top: 0;
        font-size: 20px;
    }
    .executive-summary .highlight {
        background: #7c3aed;
        color: white;
        padding: 2px 8px;
        border-radius: 4px;
        font-weight: bold;
    }
    .executive-summary .stat {
        font-weight: bold;
        color: #5b21b6;
        font-size: 16px;
    }
    </style>
""", unsafe_allow_html=True)

# Configurar estilo de seaborn
sns.set_style("whitegrid")
plt.rcParams['font.size'] = 11
plt.rcParams['axes.labelsize'] = 12
plt.rcParams['axes.titlesize'] = 14
plt.rcParams['legend.fontsize'] = 11

# Inicializar estado
if 'df' not in st.session_state:
    st.session_state.df = None
if 'df_portafolio' not in st.session_state:
    st.session_state.df_portafolio = None
if 'df_externas' not in st.session_state:
    st.session_state.df_externas = None
if 'df_filtrado' not in st.session_state:
    st.session_state.df_filtrado = None
if 'df_externas_filtrado' not in st.session_state:
    st.session_state.df_externas_filtrado = None
if 'filtros_aplicados' not in st.session_state:
    st.session_state.filtros_aplicados = False
if 'archivo_cargado' not in st.session_state:
    st.session_state.archivo_cargado = False

# Título principal
st.title("📊 Tablero resumen de gestión de Autorizaciones y Programación en Tramita")

# ======================== SECCIÓN DE CARGA (COLAPSABLE) ========================
with st.expander("📂 Cargar Archivo de Solicitudes", expanded=False):
    archivo = st.file_uploader(
        "Selecciona un archivo Excel que contenga las hojas: 'Datos', 'Portafolio' y 'Solicitudes Externas'",
        type=['xlsx', 'xls'],
        help="El archivo debe contener: Hoja 'Datos' con los campos: Tag, Solicitado, Auditado, Sede, Doc., Paciente, Edad, Genero, Diag., Entidad, Grupo Atención, Servicio, Cups, Radicación, Radicado, Autorizado, Autorización, Vence, Entregado, Servicio, Programado, Responsable, Estado, Observación, Prioridad, idOrden, idIndigo. Hoja 'Portafolio' con los campos: CUPS, codIPS, descrCodIPS, codREPS, A, UNIDAD EJECUTORA, Codigo unidad, Sede_Portafolio. Hoja 'Solicitudes Externas' con los campos: fechaRegistroFormulario, ciudad, proceso, idPaciente, nombrePaciente, entidad, servicio, cups, estado, fechaEntregaProceso, motivoCancelacion"
    )
    
    if archivo is not None:
        try:
            # Leer el archivo Excel
            excel_file = pd.ExcelFile(archivo)
            
            # Verificar que exista la hoja 'Datos'
            if 'Datos' not in excel_file.sheet_names:
                st.error("⚠️ El archivo no contiene una hoja llamada 'Datos'")
                st.info(f"📋 Hojas disponibles: {', '.join(excel_file.sheet_names)}")
                st.session_state.df = None
                st.session_state.df_portafolio = None
                st.session_state.df_externas = None
                st.session_state.archivo_cargado = False
            else:
                # Leer hoja de datos
                df = pd.read_excel(archivo, sheet_name='Datos', header=1)
                
                # Filtrar columnas no nombradas de manera segura
                unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col)]
                if unnamed_cols:
                    df = df.drop(columns=unnamed_cols)
                
                cols = df.columns.tolist()
                servicio_count = 0
                for i, col in enumerate(cols):
                    if col == 'Servicio':
                        servicio_count += 1
                        if servicio_count == 2:
                            cols[i] = 'Servicio proceso tramita'
                df.columns = cols
                
                # Guardar el nombre de las columnas
                st.session_state.header_row = df.columns.tolist()
                
                # Verificar que las columnas necesarias existan
                columnas_requeridas = ['Tag', 'Solicitado', 'Auditado', 'Sede', 'Doc.', 'Paciente', 
                                       'Edad', 'Genero', 'Diag.', 'Entidad', 'Grupo Atención', 
                                       'Servicio', 'Cups', 'Radicación', 'Radicado', 'Autorizado', 
                                       'Autorización', 'Vence', 'Entregado', 'Servicio proceso tramita', 
                                       'Programado', 'Responsable', 'Estado', 'Observación', 'Prioridad', 
                                       'idOrden', 'idIndigo']
                
                # Normalizar nombres de columnas para comparación
                columnas_df = [str(col).strip() for col in df.columns]
                columnas_requeridas_norm = [str(col).strip() for col in columnas_requeridas]
                
                # Verificar columnas faltantes
                columnas_faltantes = []
                for i, col in enumerate(columnas_requeridas_norm):
                    if col not in columnas_df:
                        if col == 'Servicio proceso tramita':
                            continue
                        columnas_faltantes.append(columnas_requeridas[i])
                
                if 'Servicio proceso tramita' in columnas_requeridas_norm:
                    servicio_count_df = sum(1 for col in columnas_df if col == 'Servicio')
                    if servicio_count_df == 1 and 'Servicio proceso tramita' in columnas_faltantes:
                        columnas_faltantes.remove('Servicio proceso tramita')
                
                if columnas_faltantes:
                    st.error(f"⚠️ La hoja 'Datos' no contiene las siguientes columnas requeridas: {', '.join(columnas_faltantes)}")
                    st.info(f"📋 Columnas encontradas: {', '.join(df.columns.tolist())}")
                    st.session_state.df = None
                    st.session_state.df_portafolio = None
                    st.session_state.df_externas = None
                    st.session_state.archivo_cargado = False
                else:
                    # Limpiar datos vacíos
                    df = df.dropna(how='all')
                    
                    # Convertir 'Solicitado' a datetime
                    df['Solicitado'] = pd.to_datetime(df['Solicitado'])
                    
                    # Convertir 'Entregado' a datetime si existe
                    if 'Entregado' in df.columns:
                        df['Entregado'] = pd.to_datetime(df['Entregado'])
                    
                    st.session_state.df = df
                    
                    # ======================== LEER PORTAFOLIO ========================
                    if 'Portafolio' not in excel_file.sheet_names:
                        st.error("⚠️ El archivo no contiene una hoja llamada 'Portafolio'")
                        st.info(f"📋 Hojas disponibles: {', '.join(excel_file.sheet_names)}")
                        st.session_state.df_portafolio = None
                        st.session_state.df_externas = None
                        st.session_state.archivo_cargado = False
                    else:
                        df_portafolio = pd.read_excel(archivo, sheet_name='Portafolio')
                        
                        columnas_portafolio = ['CUPS', 'UNIDAD EJECUTORA', 'Sede_Portafolio']
                        columnas_portafolio_faltantes = [col for col in columnas_portafolio if col not in df_portafolio.columns]
                        
                        if columnas_portafolio_faltantes:
                            st.error(f"⚠️ La hoja 'Portafolio' no contiene las siguientes columnas requeridas: {', '.join(columnas_portafolio_faltantes)}")
                            st.info(f"📋 Columnas encontradas: {', '.join(df_portafolio.columns.tolist())}")
                            st.session_state.df_portafolio = None
                            st.session_state.df_externas = None
                            st.session_state.archivo_cargado = False
                        else:
                            df_portafolio = df_portafolio.dropna(how='all')
                            st.session_state.df_portafolio = df_portafolio
                            
                            # ======================== LEER SOLICITUDES EXTERNAS ========================
                            if 'Solicitudes Externas' not in excel_file.sheet_names:
                                st.warning("⚠️ El archivo no contiene una hoja llamada 'Solicitudes Externas'. Esta hoja es opcional.")
                                st.session_state.df_externas = None
                                st.session_state.archivo_cargado = True
                            else:
                                df_externas = pd.read_excel(archivo, sheet_name='Solicitudes Externas')
                                
                                # Verificar columnas necesarias
                                columnas_externas = ['fechaRegistroFormulario', 'ciudad', 'proceso', 'idPaciente', 
                                                    'nombrePaciente', 'entidad', 'servicio', 'cups', 'estado', 
                                                    'fechaEntregaProceso', 'motivoCancelacion']
                                columnas_externas_faltantes = [col for col in columnas_externas if col not in df_externas.columns]
                                
                                if columnas_externas_faltantes:
                                    st.warning(f"⚠️ La hoja 'Solicitudes Externas' no contiene las columnas requeridas. Se omitirá.")
                                    st.info(f"📋 Columnas faltantes: {', '.join(columnas_externas_faltantes)}")
                                    st.session_state.df_externas = None
                                else:
                                    # Limpiar datos
                                    df_externas = df_externas.dropna(how='all')
                                    
                                    # Convertir fechas
                                    df_externas['fechaRegistroFormulario'] = pd.to_datetime(df_externas['fechaRegistroFormulario'], errors='coerce')
                                    df_externas['fechaEntregaProceso'] = pd.to_datetime(df_externas['fechaEntregaProceso'], errors='coerce')
                                    
                                    # Normalizar ciudad para comparación (mayúsculas y sin espacios)
                                    df_externas['ciudad_norm'] = df_externas['ciudad'].astype(str).str.strip().str.upper()
                                    # Eliminar puntos y comas
                                    df_externas['ciudad_norm'] = df_externas['ciudad_norm'].str.replace('.', '').str.replace(',', '')
                                    
                                    # Normalizar estado para comparación
                                    df_externas['estado_norm'] = df_externas['estado'].astype(str).str.strip().str.upper()
                                    
                                    st.session_state.df_externas = df_externas
                                
                                st.session_state.archivo_cargado = True
                            
                            st.session_state.filtros_aplicados = False
                            st.session_state.df_filtrado = None
                            st.session_state.df_externas_filtrado = None
                            
                            if len(df) > 0:
                                st.session_state.fecha_inicio = df['Solicitado'].min().date()
                                st.session_state.fecha_fin = df['Solicitado'].max().date()
                            
                            st.success(f"✅ Archivo cargado correctamente.")
                            st.info(f"📊 Datos: {len(df)} registros encontrados.")
                            st.info(f"📊 Portafolio: {len(df_portafolio)} registros encontrados.")
                            if st.session_state.df_externas is not None:
                                st.info(f"📊 Solicitudes Externas: {len(df_externas)} registros encontrados.")
                            st.info(f"📅 Rango de fechas: {df['Solicitado'].min().strftime('%Y-%m-%d')} - {df['Solicitado'].max().strftime('%Y-%m-%d')}")
                    
        except Exception as e:
            st.error(f"⚠️ Error al leer el archivo: {e}")
            import traceback
            st.error(f"Detalles del error: {traceback.format_exc()}")
            st.session_state.df = None
            st.session_state.df_portafolio = None
            st.session_state.df_externas = None
            st.session_state.archivo_cargado = False
    else:
        st.info("📌 Carga un archivo Excel para comenzar a trabajar")

# ======================== FUNCIÓN PARA ASIGNAR ÁREA ========================
def asignar_area_mejorada(df_data, df_portafolio):
    """
    Asigna el área a cada registro basándose en el CUPS y la Sede.
    Maneja diferentes formatos de CUPS y normaliza los nombres de sedes.
    """
    df_data_copy = df_data.copy()
    df_portafolio_copy = df_portafolio.copy()
    
    df_data_copy['Cups_clean'] = df_data_copy['Cups'].astype(str).str.strip().str[:6]
    df_portafolio_copy['CUPS_clean'] = df_portafolio_copy['CUPS'].astype(str).str.strip().str[:6]
    
    df_data_copy['Sede_clean'] = df_data_copy['Sede'].astype(str).str.strip().str.upper()
    df_portafolio_copy['Sede_Portafolio_clean'] = df_portafolio_copy['Sede_Portafolio'].astype(str).str.strip().str.upper()
    
    dict_cups_sede_area = {}
    for idx, row in df_portafolio_copy.iterrows():
        key = (row['CUPS_clean'], row['Sede_Portafolio_clean'])
        dict_cups_sede_area[key] = row['UNIDAD EJECUTORA']
    
    dict_cups_area_fallback = {}
    for idx, row in df_portafolio_copy.iterrows():
        key = row['CUPS_clean']
        if key not in dict_cups_area_fallback:
            dict_cups_area_fallback[key] = row['UNIDAD EJECUTORA']
    
    def get_area(row):
        key = (row['Cups_clean'], row['Sede_clean'])
        if key in dict_cups_sede_area:
            return dict_cups_sede_area[key]
        else:
            if row['Cups_clean'] in dict_cups_area_fallback:
                return dict_cups_area_fallback[row['Cups_clean']]
            else:
                return 'Sin Área'
    
    df_data_copy['Area'] = df_data_copy.apply(get_area, axis=1)
    df_data_copy = df_data_copy.drop(['Cups_clean', 'Sede_clean'], axis=1)
    
    return df_data_copy

# ======================== FUNCIÓN PARA CLASIFICAR GESTIÓN DE EXTERNAS ========================
def clasificar_gestion_externa(estado):
    """
    Clasifica si una solicitud externa está gestionada o pendiente.
    Solo "PENDIENTE / REGISTRADA" se considera pendiente.
    Todos los demás estados se consideran gestionados.
    """
    estado_norm = str(estado).strip().upper()
    if estado_norm == "PENDIENTE / REGISTRADA":
        return "Pendiente"
    else:
        return "Gestionado"

# ======================== FUNCIÓN PARA GENERAR RESUMEN EJECUTIVO ========================
def generar_resumen_ejecutivo(df, df_externas_filtrado):
    """Genera un resumen ejecutivo con los principales hallazgos del análisis"""
    
    total_ordenes = len(df)
    total_entidades = df['Entidad'].nunique()
    total_pacientes = df['Paciente'].nunique()
    
    estados_gestion = df['Estado_Gestion'].value_counts()
    pendientes_prog = estados_gestion.get('Pendiente gestión desde programación', 0)
    pendientes_aut = estados_gestion.get('Pendiente gestión desde Autorizaciones', 0)
    gestionados_prog = estados_gestion.get('Gestionado desde programación', 0)
    gestionados_aut = estados_gestion.get('Gestionado / En seguimiento desde Autorizaciones', 0)
    
    total_gestionados = gestionados_prog + gestionados_aut
    total_pendientes = pendientes_prog + pendientes_aut
    
    top_area = df['Area'].value_counts()
    area_top = top_area.index[0] if len(top_area) > 0 else "N/A"
    area_top_count = top_area.iloc[0] if len(top_area) > 0 else 0
    
    top_entidad = df['Entidad'].value_counts()
    entidad_top = top_entidad.index[0] if len(top_entidad) > 0 else "N/A"
    entidad_top_count = top_entidad.iloc[0] if len(top_entidad) > 0 else 0
    
    if 'dias_entrega' in df.columns:
        dias_entrega_validos = df['dias_entrega'].dropna()
        dias_entrega_validos = dias_entrega_validos[dias_entrega_validos >= 0]
        promedio_dias = dias_entrega_validos.mean() if len(dias_entrega_validos) > 0 else 0
    else:
        promedio_dias = 0
    
    # Calcular Ordenamientos/día por sede (solo días laborales)
    if 'Sede' in df.columns and 'Solicitado' in df.columns:
        df_laboral = df[df['Solicitado'].dt.weekday < 5].copy()
        if len(df_laboral) > 0:
            ordenamientos_por_dia_sede = df_laboral.groupby([df_laboral['Solicitado'].dt.date, 'Sede']).size()
            promedio_dia_sede = ordenamientos_por_dia_sede.mean() if len(ordenamientos_por_dia_sede) > 0 else 0
        else:
            promedio_dia_sede = 0
    else:
        promedio_dia_sede = 0
    
    # Calcular Ordenamientos/día por paciente
    if 'Doc.' in df.columns and 'Solicitado' in df.columns:
        ordenamientos_por_paciente_dia = df.groupby(['Doc.', df['Solicitado'].dt.date]).size()
        promedio_paciente_dia = ordenamientos_por_paciente_dia.mean() if len(ordenamientos_por_paciente_dia) > 0 else 0
    else:
        promedio_paciente_dia = 0
    
    # Construir el resumen
    resumen = '<div class="executive-summary">'
    resumen += '<h3>📋 Resumen Ejecutivo</h3>'
    resumen += f'<p><strong>Visión General:</strong> Se identificaron <span class="stat">{total_ordenes:,}</span> órdenes correspondientes a <span class="stat">{total_pacientes:,}</span> pacientes y <span class="stat">{total_entidades}</span> entidades diferentes.</p>'
    resumen += f'<p><strong>Gestión de Órdenes:</strong> Del total de órdenes, <span class="stat">{total_gestionados:,} ({total_gestionados/total_ordenes*100:.1f}%)</span> ya han sido gestionadas, mientras que <span class="stat">{total_pendientes:,} ({total_pendientes/total_ordenes*100:.1f}%)</span> se encuentran pendientes de gestión.</p>'
    
    if total_pendientes > 0:
        resumen += f'<p><strong>Cuellos de Botella:</strong> De las órdenes pendientes, <span class="stat">{pendientes_prog:,} ({pendientes_prog/total_pendientes*100:.1f}%)</span> están pendientes desde programación y <span class="stat">{pendientes_aut:,} ({pendientes_aut/total_pendientes*100:.1f}%)</span> desde autorizaciones.</p>'
    
    if area_top != "N/A":
        resumen += f'<p><strong>Concentración por Área:</strong> El área con mayor volumen de órdenes es <span class="stat">"{area_top}"</span> con <span class="stat">{area_top_count:,}</span> órdenes (<span class="stat">{area_top_count/total_ordenes*100:.1f}%</span> del total).</p>'
    
    resumen += f'<p><strong>Concentración por Entidad:</strong> La entidad con mayor participación es <span class="stat">"{entidad_top}"</span> con <span class="stat">{entidad_top_count:,}</span> órdenes (<span class="stat">{entidad_top_count/total_ordenes*100:.1f}%</span> del total).</p>'
    
    if promedio_dias > 0:
        resumen += f'<p><strong>Tiempos de Gestión:</strong> El tiempo promedio de entrega de Autorizaciones a Programación es de <span class="stat">{promedio_dias:.1f}</span> días.</p>'
    
    if promedio_dia_sede > 0:
        resumen += f'<p><strong>Productividad por Sede:</strong> En promedio se generan <span class="stat">{promedio_dia_sede:.1f}</span> órdenes por día hábil en la sede.</p>'
    
    if promedio_paciente_dia > 0:
        resumen += f'<p><strong>Productividad por Paciente:</strong> En promedio cada paciente genera <span class="stat">{promedio_paciente_dia:.1f}</span> órdenes por día.</p>'
    
    # ======================== SECCIÓN DE SOLICITUDES EXTERNAS ========================
    if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
        total_externas = len(df_externas_filtrado)
        
        # Calcular gestionados vs no gestionados usando la nueva función
        df_ext_temp = df_externas_filtrado.copy()
        df_ext_temp['gestion_clasificacion'] = df_ext_temp['estado'].apply(clasificar_gestion_externa)
        
        total_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Gestionado').sum()
        total_no_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Pendiente').sum()
        
        # Calcular días de entrega para registros con estado "ENTREGADA A PROCESO"
        entregados = df_ext_temp[df_ext_temp['estado_norm'] == 'ENTREGADA A PROCESO'].copy()
        promedio_dias_entrega_ext = None
        num_entregados_validos = 0
        
        if len(entregados) > 0:
            entregados['dias_entrega_ext'] = (entregados['fechaEntregaProceso'] - entregados['fechaRegistroFormulario']).dt.total_seconds() / (24 * 3600)
            entregados_validos = entregados[entregados['dias_entrega_ext'].notna() & (entregados['dias_entrega_ext'] >= 0)]
            num_entregados_validos = len(entregados_validos)
            if num_entregados_validos > 0:
                promedio_dias_entrega_ext = entregados_validos['dias_entrega_ext'].mean()
        
        # Construir la narrativa combinada
        resumen += f'<p><strong>Solicitudes Externas:</strong> Se identificaron <span class="stat">{total_externas:,}</span> solicitudes externas para las ciudades seleccionadas.'
        resumen += f' De estas, <span class="stat">{total_gestionados_ext:,} ({total_gestionados_ext/total_externas*100:.1f}%)</span> ya han sido gestionadas y <span class="stat">{total_no_gestionados_ext:,} ({total_no_gestionados_ext/total_externas*100:.1f}%)</span> se encuentran pendientes de gestión.'
        
        if promedio_dias_entrega_ext is not None and num_entregados_validos > 0:
            resumen += f' Para las solicitudes entregadas a proceso, el tiempo promedio de entrega es de <span class="stat">{promedio_dias_entrega_ext:.1f}</span> días, calculado sobre <span class="stat">{num_entregados_validos}</span> registros con fechas válidas.'
        elif len(entregados) > 0:
            resumen += f' No se encontraron registros con fechas válidas para calcular el tiempo promedio de entrega de las solicitudes entregadas a proceso.'
        else:
            resumen += f' No hay registros con estado "ENTREGADA A PROCESO" para calcular el tiempo promedio de entrega.'
        
        resumen += '</p>'
        
    elif df_externas_filtrado is not None and len(df_externas_filtrado) == 0:
        resumen += '<p><strong>Solicitudes Externas:</strong> No se encontraron solicitudes externas para las ciudades seleccionadas.</p>'
    else:
        resumen += '<p><strong>Solicitudes Externas:</strong> No hay datos disponibles.</p>'
    
    resumen += '</div>'
    
    return resumen

# ======================== FUNCIÓN PARA GENERAR INTERPRETACIONES ========================
def generar_interpretacion(titulo, texto):
    """Genera una interpretación con formato HTML correcto"""
    return f'<div class="interpretation-box"><strong>📝 {titulo}:</strong> {texto}</div>'

# ======================== CONTENIDO PRINCIPAL ========================
if st.session_state.archivo_cargado and st.session_state.df is not None and st.session_state.df_portafolio is not None:
    df = st.session_state.df.copy()
    df_portafolio = st.session_state.df_portafolio.copy()
    df_externas = st.session_state.df_externas.copy() if st.session_state.df_externas is not None else None
    
    # Verificar columnas datetime
    if not pd.api.types.is_datetime64_any_dtype(df['Solicitado']):
        try:
            df['Solicitado'] = pd.to_datetime(df['Solicitado'])
        except:
            st.error("⚠️ No se pudo convertir la columna 'Solicitado' a formato de fecha")
            st.stop()
    
    if 'Entregado' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Entregado']):
        try:
            df['Entregado'] = pd.to_datetime(df['Entregado'])
        except:
            pass
    
    if len(df) == 0:
        st.warning("⚠️ El archivo no contiene datos después de la fila de título")
        st.stop()
    
    # Aplicar la asignación de área
    df = asignar_area_mejorada(df, df_portafolio)
    
    def clasificar_estado_gestion(estado):
        if estado == "PROGRAMAR":
            return "Pendiente gestión desde programación"
        elif estado == "RADICAR":
            return "Pendiente gestión desde Autorizaciones"
        elif estado in ["PROGRAMADO", "PENDIENTE PROGRAMAR"]:
            return "Gestionado desde programación"
        else:
            return "Gestionado / En seguimiento desde Autorizaciones"
    
    df['Estado_Gestion'] = df['Estado'].apply(clasificar_estado_gestion)
    
    if 'Entregado' in df.columns:
        df['dias_entrega'] = (df['Entregado'] - df['Solicitado']).dt.total_seconds() / (24 * 3600)
    
    # Mostrar estadísticas de asignación para depuración
    with st.expander("📊 Estadísticas de Asignación de Áreas", expanded=False):
        conteo_areas = df['Area'].value_counts()
        total_sin_area = (df['Area'] == 'Sin Área').sum()
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total de registros", len(df))
            st.metric("Registros sin área", total_sin_area)
            st.metric("Porcentaje sin área", f"{(total_sin_area/len(df)*100):.1f}%" if len(df) > 0 else "0%")
        
        with col2:
            st.write("Distribución de áreas:")
            st.dataframe(conteo_areas.reset_index().rename(columns={'index': 'Área', 'Area': 'Cantidad'}))
        
        if total_sin_area > 0:
            st.warning("⚠️ Algunos registros no pudieron ser asignados a un área. Verifica que los CUPS y Sedes coincidan con el portafolio base.")
            ejemplos_sin_area = df[df['Area'] == 'Sin Área'][['Cups', 'Sede']].head(10)
            st.dataframe(ejemplos_sin_area)
    
    # ======================== BARRA DE FILTROS ========================
    st.markdown("### 🔍 Panel de Filtros")
    
    with st.form(key="filtros_form"):
        col_f1, col_f2, col_f3, col_f4 = st.columns([2, 2, 2, 1])
        
        with col_f1:
            fecha_min = df['Solicitado'].min().date()
            fecha_max = df['Solicitado'].max().date()
            fecha_inicio = st.date_input(
                "📅 Desde",
                value=fecha_min,
                min_value=fecha_min,
                max_value=fecha_max,
                key="fecha_inicio_dashboard"
            )
        
        with col_f2:
            fecha_fin = st.date_input(
                "📅 Hasta",
                value=fecha_max,
                min_value=fecha_min,
                max_value=fecha_max,
                key="fecha_fin_dashboard"
            )
        
        with col_f3:
            estados_disponibles = sorted(df['Estado'].dropna().unique().tolist())
            estados_seleccionados = st.multiselect(
                "📌 Estado",
                options=estados_disponibles,
                default=estados_disponibles,
                key="estados_dashboard"
            )
        
        with col_f4:
            st.write("")
            st.write("")
            aplicar_filtros = st.form_submit_button("🔍 Aplicar Filtros", use_container_width=True)
        
        col_f5, col_f6, col_f7 = st.columns([2, 2, 2])
        
        with col_f5:
            entidades_disponibles = sorted(df['Entidad'].dropna().unique().tolist())
            entidades_seleccionadas = st.multiselect(
                "🏥 Entidad",
                options=entidades_disponibles,
                default=entidades_disponibles,
                key="entidades_dashboard"
            )
        
        with col_f6:
            areas_disponibles = sorted(df['Area'].dropna().unique().tolist())
            areas_seleccionadas = st.multiselect(
                "📂 Área",
                options=areas_disponibles,
                default=areas_disponibles,
                key="areas_dashboard"
            )
        
        with col_f7:
            sedes_disponibles = sorted(df['Sede'].dropna().unique().tolist())
            sedes_seleccionadas = st.multiselect(
                "📍 Sede",
                options=sedes_disponibles,
                default=sedes_disponibles,
                key="sedes_dashboard"
            )
    
    col_reset1, col_reset2 = st.columns([1, 5])
    with col_reset1:
        if st.button("🔄 Restablecer Filtros", key="reset_filters_dashboard"):
            st.session_state.df_filtrado = None
            st.session_state.df_externas_filtrado = None
            st.session_state.filtros_aplicados = False
            st.rerun()
    
    # ======================== APLICAR FILTROS ========================
    if st.session_state.df_filtrado is None or aplicar_filtros:
        df_filtrado = df.copy()
        
        if fecha_inicio and fecha_fin:
            if fecha_inicio <= fecha_fin:
                fecha_inicio_dt = pd.Timestamp(fecha_inicio)
                fecha_fin_dt = pd.Timestamp(fecha_fin) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                df_filtrado = df_filtrado[(df_filtrado['Solicitado'] >= fecha_inicio_dt) & (df_filtrado['Solicitado'] <= fecha_fin_dt)]
            else:
                st.warning("⚠️ La fecha de inicio debe ser menor o igual a la fecha de fin")
        
        if estados_seleccionados:
            df_filtrado = df_filtrado[df_filtrado['Estado'].isin(estados_seleccionados)]
        if entidades_seleccionadas:
            df_filtrado = df_filtrado[df_filtrado['Entidad'].isin(entidades_seleccionadas)]
        if areas_seleccionadas:
            df_filtrado = df_filtrado[df_filtrado['Area'].isin(areas_seleccionadas)]
        if sedes_seleccionadas:
            df_filtrado = df_filtrado[df_filtrado['Sede'].isin(sedes_seleccionadas)]
        
        if sedes_seleccionadas:
            df_portafolio_filtrado = df_portafolio[df_portafolio['Sede_Portafolio'].isin(sedes_seleccionadas)]
        else:
            df_portafolio_filtrado = df_portafolio.copy()
        
        df_filtrado = asignar_area_mejorada(df_filtrado, df_portafolio_filtrado)
        df_filtrado['Estado_Gestion'] = df_filtrado['Estado'].apply(clasificar_estado_gestion)
        
        if 'Entregado' in df_filtrado.columns:
            df_filtrado['dias_entrega'] = (df_filtrado['Entregado'] - df_filtrado['Solicitado']).dt.total_seconds() / (24 * 3600)
        
        # ======================== FILTRAR SOLICITUDES EXTERNAS ========================
        df_externas_filtrado = None
        if df_externas is not None and len(df_externas) > 0:
            df_externas_filtrado = df_externas.copy()
            
            # Normalizar ciudades de las sedes seleccionadas
            if sedes_seleccionadas:
                # Crear lista de palabras clave para coincidencia desde la segunda palabra
                palabras_clave = []
                for sede in sedes_seleccionadas:
                    # Limpiar y normalizar la sede
                    sede_clean = str(sede).strip().upper()
                    # Eliminar puntos y comas
                    sede_clean = sede_clean.replace('.', '').replace(',', '')
                    
                    # Dividir por espacios y tomar desde la segunda palabra
                    partes = sede_clean.split()
                    if len(partes) > 1:
                        # Tomar desde la segunda palabra en adelante
                        palabra_clave = ' '.join(partes[1:])
                        if len(palabra_clave) > 0:
                            palabras_clave.append(palabra_clave)
                    else:
                        # Si solo tiene una palabra, usar toda
                        palabras_clave.append(sede_clean)
                    
                    # También agregar la sede completa como respaldo
                    palabras_clave.append(sede_clean)
                    
                    # Agregar la primera palabra también como respaldo
                    if len(partes) > 0:
                        palabras_clave.append(partes[0])
                
                # Eliminar duplicados y palabras muy cortas
                palabras_clave = list(set([p for p in palabras_clave if len(p) > 1]))
                
                # Función para determinar si la ciudad coincide con la sede (desde la segunda palabra)
                def ciudad_coincide(ciudad_norm):
                    if pd.isna(ciudad_norm):
                        return False
                    ciudad_str = str(ciudad_norm).strip().upper()
                    # Eliminar puntos y comas
                    ciudad_str = ciudad_str.replace('.', '').replace(',', '')
                    
                    for palabra in palabras_clave:
                        # Verificar si la palabra clave está contenida en la ciudad
                        if palabra in ciudad_str:
                            return True
                        # Verificar si la ciudad está contenida en la palabra clave
                        if len(palabra) > 3 and ciudad_str in palabra:
                            return True
                    return False
                
                # Aplicar el filtro
                mask = df_externas_filtrado['ciudad_norm'].apply(ciudad_coincide)
                df_externas_filtrado = df_externas_filtrado[mask]
                
                # Si no hay resultados, intentar con búsqueda más flexible
                if len(df_externas_filtrado) == 0 and len(palabras_clave) > 0:
                    # Buscar coincidencias con palabras individuales
                    for palabra in palabras_clave:
                        if len(palabra) > 2:
                            # Dividir la palabra clave en partes
                            partes_palabra = palabra.split()
                            for parte in partes_palabra:
                                if len(parte) > 2:
                                    mask = df_externas_filtrado['ciudad_norm'].str.contains(parte, na=False)
                                    if mask.any():
                                        df_externas_filtrado = df_externas_filtrado[mask]
                                        break
                            if len(df_externas_filtrado) > 0:
                                break
            
            # Filtrar por fecha si es posible
            if fecha_inicio and fecha_fin and 'fechaRegistroFormulario' in df_externas_filtrado.columns:
                fecha_inicio_dt = pd.Timestamp(fecha_inicio)
                fecha_fin_dt = pd.Timestamp(fecha_fin) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                df_externas_filtrado = df_externas_filtrado[
                    (df_externas_filtrado['fechaRegistroFormulario'] >= fecha_inicio_dt) & 
                    (df_externas_filtrado['fechaRegistroFormulario'] <= fecha_fin_dt)
                ]
            
            # Limpiar datos vacíos
            if df_externas_filtrado is not None:
                df_externas_filtrado = df_externas_filtrado.dropna(how='all')
        
        st.session_state.df_filtrado = df_filtrado
        st.session_state.df_externas_filtrado = df_externas_filtrado
        st.session_state.filtros_aplicados = True
        
        if len(df_filtrado) > 0:
            st.success(f"✅ Filtros aplicados: {len(df_filtrado)} registros encontrados")
            if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
                st.info(f"📊 Solicitudes externas filtradas: {len(df_externas_filtrado)} registros encontrados para las ciudades seleccionadas")
            elif df_externas_filtrado is not None:
                st.info("📊 Solicitudes externas filtradas: No se encontraron registros para las ciudades seleccionadas")
        else:
            st.warning("⚠️ No hay datos con los filtros seleccionados")
    
    df_filtrado = st.session_state.df_filtrado.copy()
    df_externas_filtrado = st.session_state.df_externas_filtrado.copy() if st.session_state.df_externas_filtrado is not None else None
    
    # ======================== RESUMEN EJECUTIVO ========================
    if len(df_filtrado) > 0:
        st.markdown("### 📋 Resumen Ejecutivo")
        resumen_html = generar_resumen_ejecutivo(df_filtrado, df_externas_filtrado)
        st.markdown(resumen_html, unsafe_allow_html=True)
    else:
        st.warning("⚠️ No hay datos para mostrar el resumen ejecutivo")
    
    # ======================== KPI CARDS ========================
    st.markdown("### 📊 Indicadores Clave")
    
    total_registros = len(df_filtrado)
    total_entidades = df_filtrado['Entidad'].nunique() if len(df_filtrado) > 0 else 0
    total_pacientes = df_filtrado['Paciente'].nunique() if len(df_filtrado) > 0 else 0
    
    if 'Entregado' in df_filtrado.columns and len(df_filtrado) > 0:
        dias_entrega_validos = df_filtrado['dias_entrega'].dropna()
        dias_entrega_validos = dias_entrega_validos[dias_entrega_validos >= 0]
        promedio_dias_entrega = f"{dias_entrega_validos.mean():.1f}" if len(dias_entrega_validos) > 0 else "N/A"
    else:
        promedio_dias_entrega = "N/A"
    
    if len(df_filtrado) > 0 and 'Sede' in df_filtrado.columns and 'Solicitado' in df_filtrado.columns:
        df_laboral = df_filtrado[df_filtrado['Solicitado'].dt.weekday < 5].copy()
        
        if len(df_laboral) > 0:
            ordenamientos_por_dia_sede = df_laboral.groupby([df_laboral['Solicitado'].dt.date, 'Sede']).size()
            promedio_dia_sede = f"{ordenamientos_por_dia_sede.mean():.1f}" if len(ordenamientos_por_dia_sede) > 0 else "N/A"
        else:
            promedio_dia_sede = "N/A"
    else:
        promedio_dia_sede = "N/A"
    
    if len(df_filtrado) > 0 and 'Doc.' in df_filtrado.columns:
        ordenamientos_por_paciente_dia = df_filtrado.groupby(['Doc.', df_filtrado['Solicitado'].dt.date]).size()
        promedio_paciente_dia = f"{ordenamientos_por_paciente_dia.mean():.1f}" if len(ordenamientos_por_paciente_dia) > 0 else "N/A"
    else:
        promedio_paciente_dia = "N/A"
    
    # ======================== MÉTRICAS DE SOLICITUDES EXTERNAS ========================
    total_externas = 0
    total_gestionados_ext = 0
    pct_gestionados_ext = 0
    promedio_dias_entrega_ext = None
    num_entregados_validos = 0
    
    if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
        total_externas = len(df_externas_filtrado)
        df_ext_temp = df_externas_filtrado.copy()
        df_ext_temp['gestion_clasificacion'] = df_ext_temp['estado'].apply(clasificar_gestion_externa)
        total_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Gestionado').sum()
        pct_gestionados_ext = (total_gestionados_ext / total_externas * 100) if total_externas > 0 else 0
        
        # Calcular días de entrega para entregados a proceso
        entregados = df_ext_temp[df_ext_temp['estado_norm'] == 'ENTREGADA A PROCESO'].copy()
        if len(entregados) > 0:
            entregados['dias_entrega_ext'] = (entregados['fechaEntregaProceso'] - entregados['fechaRegistroFormulario']).dt.total_seconds() / (24 * 3600)
            entregados_validos = entregados[entregados['dias_entrega_ext'].notna() & (entregados['dias_entrega_ext'] >= 0)]
            num_entregados_validos = len(entregados_validos)
            if num_entregados_validos > 0:
                promedio_dias_entrega_ext = entregados_validos['dias_entrega_ext'].mean()
    
    # Mostrar KPI Cards (6 tarjetas)
    if len(df_filtrado) > 0:
        # Primera fila - 3 tarjetas
        col_k1, col_k2, col_k3 = st.columns(3)
        
        with col_k1:
            st.markdown(f"""
                <div class="metric-card">
                    <p class="metric-label">📊 Total Registros</p>
                    <p class="metric-value">{total_registros:,}</p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_k2:
            st.markdown(f"""
                <div class="metric-card">
                    <p class="metric-label">🏥 Entidades</p>
                    <p class="metric-value">{total_entidades:,}</p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_k3:
            st.markdown(f"""
                <div class="metric-card">
                    <p class="metric-label">👥 Pacientes</p>
                    <p class="metric-value">{total_pacientes:,}</p>
                </div>
            """, unsafe_allow_html=True)
        
        # Segunda fila - 3 tarjetas (incluyendo las nuevas métricas de externas)
        col_k4, col_k5, col_k6 = st.columns(3)
        
        with col_k4:
            st.markdown(f"""
                <div class="metric-card-small">
                    <p class="metric-label">⏱️ Días promedio entrega</p>
                    <p class="metric-value">{promedio_dias_entrega}</p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_k5:
            st.markdown(f"""
                <div class="metric-card-small">
                    <p class="metric-label">📅 Ordenamientos/día por sede (Lun-Vie)</p>
                    <p class="metric-value">{promedio_dia_sede}</p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_k6:
            st.markdown(f"""
                <div class="metric-card-small">
                    <p class="metric-label">👤 Ordenamientos/día por paciente</p>
                    <p class="metric-value">{promedio_paciente_dia}</p>
                </div>
            """, unsafe_allow_html=True)
        
        # Tercera fila - 3 tarjetas para métricas de externas
        col_k7, col_k8, col_k9 = st.columns(3)
        
        with col_k7:
            st.markdown(f"""
                <div class="metric-card-small">
                    <p class="metric-label">📋 Solicitudes Externas</p>
                    <p class="metric-value">{total_externas:,}</p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_k8:
            if total_externas > 0:
                st.markdown(f"""
                    <div class="metric-card-small">
                        <p class="metric-label">✅ % Gestionado (Externas)</p>
                        <p class="metric-value">{pct_gestionados_ext:.1f}%</p>
                    </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                    <div class="metric-card-small">
                        <p class="metric-label">✅ % Gestionado (Externas)</p>
                        <p class="metric-value">N/A</p>
                    </div>
                """, unsafe_allow_html=True)
        
        with col_k9:
            if promedio_dias_entrega_ext is not None and num_entregados_validos > 0:
                st.markdown(f"""
                    <div class="metric-card-small">
                        <p class="metric-label">⏱️ Días entrega (Externas)</p>
                        <p class="metric-value">{promedio_dias_entrega_ext:.1f}</p>
                    </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                    <div class="metric-card-small">
                        <p class="metric-label">⏱️ Días entrega (Externas)</p>
                        <p class="metric-value">N/A</p>
                    </div>
                """, unsafe_allow_html=True)
    
    # ======================== TABLA DE RESULTADOS ========================
    with st.expander("📋 Ver Detalle de Resultados (Datos Filtrados)", expanded=False):
        st.markdown("#### Detalle de órdenes con filtros aplicados")
        
        columnas_tabla = ['Estado', 'Estado_Gestion', 'Solicitado', 'Doc.', 'Paciente', 'Entidad', 'Area', 'Cups', 'Servicio', 'Observación']
        columnas_existentes = [col for col in columnas_tabla if col in df_filtrado.columns]
        
        if columnas_existentes:
            df_tabla = df_filtrado[columnas_existentes].copy()
            
            if 'Solicitado' in df_tabla.columns:
                df_tabla['Solicitado'] = df_tabla['Solicitado'].dt.strftime('%Y-%m-%d %H:%M')
            
            st.dataframe(
                df_tabla,
                use_container_width=True,
                height=400,
                column_config={
                    "Estado": st.column_config.TextColumn("Estado", width="medium"),
                    "Estado_Gestion": st.column_config.TextColumn("Estado de Gestión", width="large"),
                    "Solicitado": st.column_config.TextColumn("Solicitado", width="medium"),
                    "Doc.": st.column_config.TextColumn("Documento", width="small"),
                    "Paciente": st.column_config.TextColumn("Paciente", width="large"),
                    "Entidad": st.column_config.TextColumn("Entidad", width="large"),
                    "Area": st.column_config.TextColumn("Área", width="large"),
                    "Cups": st.column_config.TextColumn("Cups", width="small"),
                    "Servicio": st.column_config.TextColumn("Servicio", width="medium"),
                    "Observación": st.column_config.TextColumn("Observación", width="large"),
                }
            )
            st.caption(f"📊 Mostrando {len(df_tabla)} registros")
        else:
            st.warning("No se encontraron las columnas necesarias para mostrar la tabla")
        
        # ======================== TABLA DE SOLICITUDES EXTERNAS ========================
        if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
            st.markdown("---")
            st.markdown("#### 📋 Resumen de Solicitudes Externas")
            
            # Mostrar tabla con datos relevantes
            columnas_ext = ['fechaRegistroFormulario', 'ciudad', 'proceso', 'idPaciente', 'nombrePaciente', 'entidad', 'estado']
            columnas_ext_existentes = [col for col in columnas_ext if col in df_externas_filtrado.columns]
            
            if columnas_ext_existentes:
                df_ext_tabla = df_externas_filtrado[columnas_ext_existentes].copy()
                
                # Formatear fechas
                if 'fechaRegistroFormulario' in df_ext_tabla.columns:
                    df_ext_tabla['fechaRegistroFormulario'] = pd.to_datetime(df_ext_tabla['fechaRegistroFormulario']).dt.strftime('%Y-%m-%d')
                
                # Agregar columna de clasificación
                df_ext_tabla['Clasificación'] = df_ext_tabla['estado'].apply(clasificar_gestion_externa)
                
                st.dataframe(
                    df_ext_tabla,
                    use_container_width=True,
                    height=300,
                    column_config={
                        "fechaRegistroFormulario": st.column_config.TextColumn("Fecha Registro", width="medium"),
                        "ciudad": st.column_config.TextColumn("Ciudad", width="medium"),
                        "proceso": st.column_config.TextColumn("Proceso", width="large"),
                        "idPaciente": st.column_config.TextColumn("ID Paciente", width="small"),
                        "nombrePaciente": st.column_config.TextColumn("Paciente", width="large"),
                        "entidad": st.column_config.TextColumn("Entidad", width="large"),
                        "estado": st.column_config.TextColumn("Estado", width="medium"),
                        "Clasificación": st.column_config.TextColumn("Clasificación", width="medium"),
                    }
                )
                st.caption(f"📊 Mostrando {len(df_ext_tabla)} registros de solicitudes externas")
            
            # Mostrar estadísticas adicionales
            total_externas = len(df_externas_filtrado)
            
            # Calcular gestionados vs no gestionados usando la nueva función
            df_ext_temp = df_externas_filtrado.copy()
            df_ext_temp['gestion_clasificacion'] = df_ext_temp['estado'].apply(clasificar_gestion_externa)
            
            total_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Gestionado').sum()
            total_no_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Pendiente').sum()
            
            col_ext1, col_ext2, col_ext3, col_ext4 = st.columns(4)
            with col_ext1:
                st.metric("Total Solicitudes Externas", f"{total_externas:,}")
            with col_ext2:
                st.metric("Gestionadas", f"{total_gestionados_ext:,}", delta=f"{total_gestionados_ext/total_externas*100:.1f}%")
            with col_ext3:
                st.metric("Pendientes", f"{total_no_gestionados_ext:,}", delta=f"{total_no_gestionados_ext/total_externas*100:.1f}%")
            with col_ext4:
                # Calcular días de entrega promedio para entregados
                entregados = df_ext_temp[df_ext_temp['estado_norm'] == 'ENTREGADA A PROCESO'].copy()
                if len(entregados) > 0:
                    entregados['dias_entrega_ext'] = (entregados['fechaEntregaProceso'] - entregados['fechaRegistroFormulario']).dt.total_seconds() / (24 * 3600)
                    entregados_validos = entregados[entregados['dias_entrega_ext'].notna() & (entregados['dias_entrega_ext'] >= 0)]
                    if len(entregados_validos) > 0:
                        st.metric("Promedio Días Entrega", f"{entregados_validos['dias_entrega_ext'].mean():.1f}", 
                                 delta=f"({len(entregados_validos)} registros)")
                    else:
                        st.metric("Promedio Días Entrega", "N/A")
                else:
                    st.metric("Promedio Días Entrega", "N/A")
    
    # ======================== GRÁFICOS ========================
    if len(df_filtrado) > 0:
        st.markdown("### 📈 Análisis Visual")
        
        # Paleta de colores para gráficos de anillo - Colores diferenciados y vibrantes
        colores_diferenciados = [
            '#FF6B6B',  # Rojo
            '#4ECDC4',  # Verde azulado
            '#45B7D1',  # Azul
            '#96CEB4',  # Verde claro
            '#FFEAA7',  # Amarillo
            '#DDA0DD',  # Ciruela claro
            '#FF8C94',  # Rosa
            '#A8E6CF',  # Menta
            '#D4A5A5',  # Rosa viejo
            '#9B59B6'   # Morado
        ]
        
        # ======================== GRÁFICO 1 ========================
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.subheader("📊 Ordenes Generadas vs Gestionadas")
        
        agrupacion = st.radio(
            "Agrupar por:",
            options=["Día", "Semana", "Quincena", "Mes"],
            horizontal=True,
            key="agrupacion_grafico1_dashboard"
        )
        
        df_temp = df_filtrado.copy()
        
        if agrupacion == "Día":
            df_temp['Fecha_Agrupada'] = df_temp['Solicitado'].dt.date
        elif agrupacion == "Semana":
            df_temp['Fecha_Agrupada'] = df_temp['Solicitado'].dt.to_period('W').dt.start_time
        elif agrupacion == "Quincena":
            df_temp['Dia'] = df_temp['Solicitado'].dt.day
            df_temp['Quincena'] = df_temp['Dia'].apply(lambda x: 1 if x <= 15 else 2)
            df_temp['Fecha_Agrupada'] = df_temp['Solicitado'].dt.to_period('M').dt.start_time
            df_temp['Fecha_Agrupada'] = df_temp.apply(
                lambda row: row['Fecha_Agrupada'] + pd.Timedelta(days=(row['Quincena']-1)*15), 
                axis=1
            )
        elif agrupacion == "Mes":
            df_temp['Fecha_Agrupada'] = df_temp['Solicitado'].dt.to_period('M').dt.start_time
        
        ordenes_generadas = df_temp.groupby('Fecha_Agrupada').size().reset_index()
        ordenes_generadas.columns = ['Fecha', 'Generadas']
        
        df_gestionadas = df_temp[df_temp['Estado'] != "RADICAR"]
        ordenes_gestionadas = df_gestionadas.groupby('Fecha_Agrupada').size().reset_index()
        ordenes_gestionadas.columns = ['Fecha', 'Gestionadas']
        
        df_graf1 = pd.merge(ordenes_generadas, ordenes_gestionadas, on='Fecha', how='outer').fillna(0)
        df_graf1 = df_graf1.sort_values('Fecha')
        
        fig1, ax1 = plt.subplots(figsize=(12, 5))
        x = np.arange(len(df_graf1['Fecha']))
        width = 0.35
        
        ax1.bar(x - width/2, df_graf1['Generadas'], width, label='Generadas', color='#6d28d9')
        ax1.bar(x + width/2, df_graf1['Gestionadas'], width, label='Gestionadas', color='#a78bfa')
        
        ax1.set_xlabel('Fecha')
        ax1.set_ylabel('Cantidad')
        ax1.set_title('Órdenes Generadas vs Gestionadas')
        ax1.set_xticks(x)
        ax1.set_xticklabels(df_graf1['Fecha'], rotation=45, ha='right', fontsize=9)
        ax1.legend()
        
        for i, v in enumerate(df_graf1['Generadas']):
            ax1.text(i - width/2, v + 0.5, str(int(v)), ha='center', va='bottom', fontsize=10, fontweight='bold', color='black')
        for i, v in enumerate(df_graf1['Gestionadas']):
            ax1.text(i + width/2, v + 0.5, str(int(v)), ha='center', va='bottom', fontsize=10, fontweight='bold', color='black')
        
        plt.tight_layout()
        st.pyplot(fig1)
        
        # Interpretación GRÁFICO 1
        total_generadas = df_graf1['Generadas'].sum()
        total_gestionadas = df_graf1['Gestionadas'].sum()
        pct_gestionadas = (total_gestionadas / total_generadas * 100) if total_generadas > 0 else 0
        pct_pendientes = 100 - pct_gestionadas
        dia_pico = df_graf1.loc[df_graf1['Generadas'].idxmax(), 'Fecha'] if len(df_graf1) > 0 else "N/A"
        max_generadas = int(df_graf1['Generadas'].max()) if len(df_graf1) > 0 else 0
        
        texto_interpretacion1 = f'Se generaron <strong>{int(total_generadas)}</strong> órdenes en total, de las cuales <strong>{int(total_gestionadas)} (<span class="stat">{pct_gestionadas:.1f}%</span>)</strong> ya fueron gestionadas y <strong>{int(total_generadas - total_gestionadas)} (<span class="stat">{pct_pendientes:.1f}%</span>)</strong> aún se encuentran pendientes de gestión. El día con mayor actividad fue <strong>{dia_pico}</strong> con <strong>{max_generadas}</strong> órdenes generadas.'
        
        st.markdown(generar_interpretacion("Interpretación", texto_interpretacion1), unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICO 2 ========================
        with st.container():
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.subheader("📊 Gestión de autorizaciones y ordenes disponibles para programación")
            
            estado_gestion_counts = df_filtrado['Estado_Gestion'].value_counts().reset_index()
            estado_gestion_counts.columns = ['Estado', 'Cantidad']
            
            colores_estados = {
                'Pendiente gestión desde programación': '#FF6B6B',
                'Pendiente gestión desde Autorizaciones': '#4ECDC4',
                'Gestionado desde programación': '#45B7D1',
                'Gestionado / En seguimiento desde Autorizaciones': '#96CEB4'
            }
            
            colors = [colores_estados.get(estado, '#CCCCCC') for estado in estado_gestion_counts['Estado']]
            
            fig2, ax2 = plt.subplots(figsize=(14, 8))
            
            wedges, texts, autotexts = ax2.pie(
                estado_gestion_counts['Cantidad'],
                labels=None,
                autopct=lambda pct: f'{pct:.1f}%',
                colors=colors,
                startangle=90,
                wedgeprops={'width': 0.4, 'edgecolor': 'white', 'linewidth': 2},
                pctdistance=0.75,
                textprops={'fontsize': 12, 'fontweight': 'bold', 'color': 'black'}
            )
            
            for autotext in autotexts:
                autotext.set_color('black')
                autotext.set_fontsize(13)
                autotext.set_fontweight('bold')
                autotext.set_bbox(dict(
                    boxstyle="round,pad=0.3", 
                    facecolor='white', 
                    edgecolor='gray', 
                    alpha=0.85,
                    linewidth=1
                ))
            
            for i, wedge in enumerate(wedges):
                ang = (wedge.theta2 + wedge.theta1) / 2
                x = 1.35 * np.cos(np.radians(ang))
                y = 1.35 * np.sin(np.radians(ang))
                
                x_mid = 1.05 * np.cos(np.radians(ang))
                y_mid = 1.05 * np.sin(np.radians(ang))
                
                ax2.plot([x_mid, x], [y_mid, y], color='gray', linewidth=1.5, linestyle='-', alpha=0.7)
                
                cantidad = estado_gestion_counts['Cantidad'].iloc[i]
                ax2.text(x, y, f"{cantidad}", 
                        fontsize=14, fontweight='bold', ha='center', va='center', 
                        color='black',
                        bbox=dict(
                            boxstyle="round,pad=0.3", 
                            facecolor='white', 
                            edgecolor='gray', 
                            alpha=0.9,
                            linewidth=1
                        ))
            
            legend_elements = []
            for i, estado in enumerate(estado_gestion_counts['Estado']):
                legend_elements.append(
                    Patch(facecolor=colors[i], edgecolor='white', linewidth=2, 
                          label=f"{estado} ({estado_gestion_counts['Cantidad'].iloc[i]})")
                )
            
            ax2.legend(
                handles=legend_elements,
                loc='center left',
                bbox_to_anchor=(1.05, 0.5),
                fontsize=11,
                title="Estados de Gestión",
                title_fontsize=13,
                framealpha=0.95,
                edgecolor='#7c3aed',
                facecolor='white',
                shadow=True,
                borderpad=1
            )
            
            ax2.set_title('Gestión de autorizaciones y ordenes disponibles para programación', 
                          fontsize=14, fontweight='bold', pad=20)
            plt.tight_layout()
            st.pyplot(fig2)
            
            # Interpretación GRÁFICO 2
            total_estados = estado_gestion_counts['Cantidad'].sum()
            estado_pendientes_prog = estado_gestion_counts[estado_gestion_counts['Estado'] == 'Pendiente gestión desde programación']['Cantidad'].sum() if 'Pendiente gestión desde programación' in estado_gestion_counts['Estado'].values else 0
            estado_pendientes_aut = estado_gestion_counts[estado_gestion_counts['Estado'] == 'Pendiente gestión desde Autorizaciones']['Cantidad'].sum() if 'Pendiente gestión desde Autorizaciones' in estado_gestion_counts['Estado'].values else 0
            estado_gestionados_prog = estado_gestion_counts[estado_gestion_counts['Estado'] == 'Gestionado desde programación']['Cantidad'].sum() if 'Gestionado desde programación' in estado_gestion_counts['Estado'].values else 0
            estado_gestionados_aut = estado_gestion_counts[estado_gestion_counts['Estado'] == 'Gestionado / En seguimiento desde Autorizaciones']['Cantidad'].sum() if 'Gestionado / En seguimiento desde Autorizaciones' in estado_gestion_counts['Estado'].values else 0
            
            mayor_carga = "Pendiente gestión desde programación" if estado_pendientes_prog == max([estado_pendientes_prog, estado_pendientes_aut, estado_gestionados_prog, estado_gestionados_aut]) else "Pendiente gestión desde Autorizaciones"
            
            texto_interpretacion2 = f'Del total de <strong>{total_estados}</strong> órdenes, <strong>{estado_pendientes_prog} (<span class="stat">{estado_pendientes_prog/total_estados*100:.1f}%</span>)</strong> se encuentran pendientes de gestión desde programación, <strong>{estado_pendientes_aut} (<span class="stat">{estado_pendientes_aut/total_estados*100:.1f}%</span>)</strong> pendientes desde autorizaciones, <strong>{estado_gestionados_prog} (<span class="stat">{estado_gestionados_prog/total_estados*100:.1f}%</span>)</strong> ya gestionadas desde programación, y <strong>{estado_gestionados_aut} (<span class="stat">{estado_gestionados_aut/total_estados*100:.1f}%</span>)</strong> gestionadas o en seguimiento desde autorizaciones. La mayor carga de trabajo se concentra en <strong>{mayor_carga}</strong>.'
            
            st.markdown(generar_interpretacion("Interpretación", texto_interpretacion2), unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICO 3 ========================
        with st.container():
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.subheader("📊 Pendientes de Gestión por Área")
            
            df_pendientes = df_filtrado[df_filtrado['Estado_Gestion'] == "Pendiente gestión desde programación"]
            
            if len(df_pendientes) > 0:
                pendientes_por_area = df_pendientes['Area'].value_counts().reset_index()
                pendientes_por_area.columns = ['Área', 'Cantidad']
                
                num_areas = len(pendientes_por_area)
                colors_area = colores_diferenciados[:num_areas]
                
                fig3, ax3 = plt.subplots(figsize=(14, 8))
                
                wedges, texts, autotexts = ax3.pie(
                    pendientes_por_area['Cantidad'],
                    labels=None,
                    autopct=lambda pct: f'{pct:.1f}%',
                    colors=colors_area,
                    startangle=90,
                    wedgeprops={'width': 0.4, 'edgecolor': 'white', 'linewidth': 2},
                    pctdistance=0.75,
                    textprops={'fontsize': 12, 'fontweight': 'bold', 'color': 'black'}
                )
                
                for autotext in autotexts:
                    autotext.set_color('black')
                    autotext.set_fontsize(12)
                    autotext.set_fontweight('bold')
                    autotext.set_bbox(dict(
                        boxstyle="round,pad=0.3", 
                        facecolor='white', 
                        edgecolor='gray', 
                        alpha=0.85,
                        linewidth=1
                    ))
                
                for i, wedge in enumerate(wedges):
                    ang = (wedge.theta2 + wedge.theta1) / 2
                    x = 1.35 * np.cos(np.radians(ang))
                    y = 1.35 * np.sin(np.radians(ang))
                    
                    x_mid = 1.05 * np.cos(np.radians(ang))
                    y_mid = 1.05 * np.sin(np.radians(ang))
                    
                    ax3.plot([x_mid, x], [y_mid, y], color='gray', linewidth=1.5, linestyle='-', alpha=0.7)
                    
                    cantidad = pendientes_por_area['Cantidad'].iloc[i]
                    ax3.text(x, y, f"{cantidad}", 
                            fontsize=13, fontweight='bold', ha='center', va='center', 
                            color='black',
                            bbox=dict(
                                boxstyle="round,pad=0.3", 
                                facecolor='white', 
                                edgecolor='gray', 
                                alpha=0.9,
                                linewidth=1
                            ))
                
                legend_elements_area = []
                for i, area in enumerate(pendientes_por_area['Área']):
                    legend_elements_area.append(
                        Patch(facecolor=colors_area[i], edgecolor='white', linewidth=2, 
                              label=f"{area} ({pendientes_por_area['Cantidad'].iloc[i]})")
                    )
                
                ax3.legend(
                    handles=legend_elements_area,
                    loc='center left',
                    bbox_to_anchor=(1.05, 0.5),
                    fontsize=11,
                    title="Áreas",
                    title_fontsize=13,
                    framealpha=0.95,
                    edgecolor='#7c3aed',
                    facecolor='white',
                    shadow=True,
                    borderpad=1
                )
                
                ax3.set_title('Pendientes de Gestión por Área', fontsize=14, fontweight='bold', pad=20)
                plt.tight_layout()
                st.pyplot(fig3)
                
                # Interpretación GRÁFICO 3
                total_pendientes = pendientes_por_area['Cantidad'].sum()
                
                if len(pendientes_por_area) > 0:
                    max_area = pendientes_por_area.iloc[0]['Área']
                    max_cantidad = pendientes_por_area.iloc[0]['Cantidad']
                    
                    texto_interpretacion3 = f'Hay <strong>{total_pendientes}</strong> órdenes pendientes de gestión desde programación, distribuidas en <strong>{len(pendientes_por_area)}</strong> áreas. El área con mayor volumen de pendientes es <span class="stat">"{max_area}"</span> con <strong>{max_cantidad}</strong> órdenes (<span class="stat">{max_cantidad/total_pendientes*100:.1f}%</span> del total).'
                    
                    if len(pendientes_por_area) > 1:
                        segunda_area = pendientes_por_area.iloc[1]['Área']
                        segunda_cantidad = pendientes_por_area.iloc[1]['Cantidad']
                        texto_interpretacion3 += f' {segunda_area} es la segunda área con <strong>{segunda_cantidad}</strong> órdenes pendientes (<span class="stat">{segunda_cantidad/total_pendientes*100:.1f}%</span> del total).'
                    
                    st.markdown(generar_interpretacion("Interpretación", texto_interpretacion3), unsafe_allow_html=True)
            else:
                st.info("No hay ordenamientos pendientes de gestión desde programación")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICO 4: Solicitudes Externas por Estado ========================
        if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
            with st.container():
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.subheader("📊 Gestión de órdenes externas por Estado")
                
                # Agrupar por estado
                externas_por_estado = df_externas_filtrado['estado'].value_counts().reset_index()
                externas_por_estado.columns = ['Estado', 'Cantidad']
                
                # Usar colores diferenciados
                num_colores_ext = len(externas_por_estado)
                colors_ext = colores_diferenciados[:num_colores_ext]
                
                fig4, ax4 = plt.subplots(figsize=(14, 8))
                
                wedges, texts, autotexts = ax4.pie(
                    externas_por_estado['Cantidad'],
                    labels=None,
                    autopct=lambda pct: f'{pct:.1f}%',
                    colors=colors_ext,
                    startangle=90,
                    wedgeprops={'width': 0.4, 'edgecolor': 'white', 'linewidth': 2},
                    pctdistance=0.75,
                    textprops={'fontsize': 12, 'fontweight': 'bold', 'color': 'black'}
                )
                
                for autotext in autotexts:
                    autotext.set_color('black')
                    autotext.set_fontsize(12)
                    autotext.set_fontweight('bold')
                    autotext.set_bbox(dict(
                        boxstyle="round,pad=0.3", 
                        facecolor='white', 
                        edgecolor='gray', 
                        alpha=0.85,
                        linewidth=1
                    ))
                
                for i, wedge in enumerate(wedges):
                    ang = (wedge.theta2 + wedge.theta1) / 2
                    x = 1.35 * np.cos(np.radians(ang))
                    y = 1.35 * np.sin(np.radians(ang))
                    
                    x_mid = 1.05 * np.cos(np.radians(ang))
                    y_mid = 1.05 * np.sin(np.radians(ang))
                    
                    ax4.plot([x_mid, x], [y_mid, y], color='gray', linewidth=1.5, linestyle='-', alpha=0.7)
                    
                    cantidad = externas_por_estado['Cantidad'].iloc[i]
                    ax4.text(x, y, f"{int(cantidad)}", 
                            fontsize=13, fontweight='bold', ha='center', va='center', 
                            color='black',
                            bbox=dict(
                                boxstyle="round,pad=0.3", 
                                facecolor='white', 
                                edgecolor='gray', 
                                alpha=0.9,
                                linewidth=1
                            ))
                
                legend_elements_ext = []
                for i, estado in enumerate(externas_por_estado['Estado']):
                    legend_elements_ext.append(
                        Patch(facecolor=colors_ext[i], edgecolor='white', linewidth=2, 
                              label=f"{estado} ({int(externas_por_estado['Cantidad'].iloc[i])})")
                    )
                
                ax4.legend(
                    handles=legend_elements_ext,
                    loc='center left',
                    bbox_to_anchor=(1.05, 0.5),
                    fontsize=11,
                    title="Estados de Solicitudes Externas",
                    title_fontsize=13,
                    framealpha=0.95,
                    edgecolor='#7c3aed',
                    facecolor='white',
                    shadow=True,
                    borderpad=1
                )
                
                ax4.set_title('Gestión de órdenes externas por Estado', fontsize=14, fontweight='bold', pad=20)
                plt.tight_layout()
                st.pyplot(fig4)
                
                # Interpretación GRÁFICO 4
                total_externas = externas_por_estado['Cantidad'].sum()
                
                if len(externas_por_estado) > 0:
                    max_estado = externas_por_estado.iloc[0]['Estado']
                    max_cantidad = externas_por_estado.iloc[0]['Cantidad']
                    
                    # Calcular gestionados vs no gestionados usando la nueva función
                    df_ext_temp = df_externas_filtrado.copy()
                    df_ext_temp['gestion_clasificacion'] = df_ext_temp['estado'].apply(clasificar_gestion_externa)
                    
                    total_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Gestionado').sum()
                    total_no_gestionados_ext = (df_ext_temp['gestion_clasificacion'] == 'Pendiente').sum()
                    
                    texto_interpretacion4 = f'Se registraron <strong>{int(total_externas)}</strong> solicitudes externas en total, distribuidas en <strong>{len(externas_por_estado)}</strong> estados. El estado con mayor volumen es <span class="stat">"{max_estado}"</span> con <strong>{int(max_cantidad)}</strong> solicitudes (<span class="stat">{max_cantidad/total_externas*100:.1f}%</span> del total).'
                    
                    if len(externas_por_estado) > 1:
                        segundo_estado = externas_por_estado.iloc[1]['Estado']
                        segunda_cantidad = externas_por_estado.iloc[1]['Cantidad']
                        texto_interpretacion4 += f' {segundo_estado} es el segundo estado con <strong>{int(segunda_cantidad)}</strong> solicitudes (<span class="stat">{segunda_cantidad/total_externas*100:.1f}%</span> del total).'
                    
                    texto_interpretacion4 += f' De estas, <strong>{total_gestionados_ext} (<span class="stat">{total_gestionados_ext/total_externas*100:.1f}%</span>)</strong> ya han sido gestionadas y <strong>{total_no_gestionados_ext} (<span class="stat">{total_no_gestionados_ext/total_externas*100:.1f}%</span>)</strong> se encuentran pendientes de gestión.'
                    
                    # Calcular días de entrega para entregados a proceso
                    entregados_ext = df_ext_temp[df_ext_temp['estado_norm'] == 'ENTREGADA A PROCESO'].copy()
                    if len(entregados_ext) > 0:
                        entregados_ext['dias_entrega_ext'] = (entregados_ext['fechaEntregaProceso'] - entregados_ext['fechaRegistroFormulario']).dt.total_seconds() / (24 * 3600)
                        entregados_validos = entregados_ext[entregados_ext['dias_entrega_ext'].notna() & (entregados_ext['dias_entrega_ext'] >= 0)]
                        if len(entregados_validos) > 0:
                            texto_interpretacion4 += f' Para las solicitudes entregadas a proceso, el tiempo promedio de entrega es de <span class="stat">{entregados_validos["dias_entrega_ext"].mean():.1f}</span> días, calculado sobre <strong>{len(entregados_validos)}</strong> registros con fechas válidas.'
                        else:
                            texto_interpretacion4 += f' No se encontraron registros con fechas válidas para calcular el tiempo promedio de entrega de las solicitudes entregadas a proceso.'
                    
                    st.markdown(generar_interpretacion("Interpretación", texto_interpretacion4), unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)
        else:
            with st.container():
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.subheader("📊 Gestión de órdenes externas por Estado")
                if df_externas is not None and len(df_externas) > 0:
                    st.info("No se encontraron solicitudes externas para las ciudades seleccionadas. Verifica que los nombres de las ciudades coincidan con las sedes.")
                else:
                    st.info("No hay datos de solicitudes externas disponibles.")
                st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICO 5: Solicitudes Externas por Proceso ========================
        if df_externas_filtrado is not None and len(df_externas_filtrado) > 0 and 'proceso' in df_externas_filtrado.columns:
            with st.container():
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.subheader("📊 Solicitudes Externas por Proceso")
                
                # Agrupar por proceso
                externas_por_proceso = df_externas_filtrado['proceso'].value_counts().reset_index()
                externas_por_proceso.columns = ['Proceso', 'Cantidad']
                externas_por_proceso = externas_por_proceso.sort_values('Cantidad', ascending=False)
                
                if len(externas_por_proceso) > 0:
                    fig5, ax5 = plt.subplots(figsize=(12, 6))
                    bars5 = ax5.bar(externas_por_proceso['Proceso'], externas_por_proceso['Cantidad'], color='#45B7D1')
                    
                    ax5.set_xlabel('Proceso')
                    ax5.set_ylabel('Cantidad de Solicitudes')
                    ax5.set_title('Solicitudes Externas por Proceso', fontsize=14, fontweight='bold')
                    
                    # Rotar etiquetas si son muchas
                    if len(externas_por_proceso) > 5:
                        ax5.set_xticklabels(externas_por_proceso['Proceso'], rotation=45, ha='right', fontsize=9)
                    else:
                        ax5.set_xticklabels(externas_por_proceso['Proceso'], rotation=0, ha='center', fontsize=10)
                    
                    # Agregar etiquetas de datos
                    for bar in bars5:
                        height = bar.get_height()
                        ax5.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                                f'{int(height)}', ha='center', va='bottom', fontsize=11, fontweight='bold', color='black')
                    
                    plt.tight_layout()
                    st.pyplot(fig5)
                    
                    # Interpretación GRÁFICO 5
                    total_procesos = externas_por_proceso['Cantidad'].sum()
                    if len(externas_por_proceso) > 0:
                        top_proceso = externas_por_proceso.iloc[0]['Proceso']
                        top_cantidad = externas_por_proceso.iloc[0]['Cantidad']
                        
                        texto_interpretacion5 = f'Se registraron <strong>{int(total_procesos)}</strong> solicitudes externas distribuidas en <strong>{len(externas_por_proceso)}</strong> procesos. El proceso con mayor volumen es <span class="stat">"{top_proceso}"</span> con <strong>{int(top_cantidad)}</strong> solicitudes (<span class="stat">{top_cantidad/total_procesos*100:.1f}%</span> del total).'
                        
                        if len(externas_por_proceso) > 1:
                            segundo_proceso = externas_por_proceso.iloc[1]['Proceso']
                            segunda_cantidad = externas_por_proceso.iloc[1]['Cantidad']
                            texto_interpretacion5 += f' {segundo_proceso} es el segundo proceso con <strong>{int(segunda_cantidad)}</strong> solicitudes (<span class="stat">{segunda_cantidad/total_procesos*100:.1f}%</span> del total).'
                        
                        st.markdown(generar_interpretacion("Interpretación", texto_interpretacion5), unsafe_allow_html=True)
                else:
                    st.info("No hay datos de procesos para mostrar")
                st.markdown('</div>', unsafe_allow_html=True)
        else:
            with st.container():
                st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                st.subheader("📊 Solicitudes Externas por Proceso")
                if df_externas is not None and len(df_externas) > 0:
                    st.info("No se encontraron solicitudes externas para las ciudades seleccionadas.")
                else:
                    st.info("No hay datos de solicitudes externas disponibles.")
                st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICOS EN DOS COLUMNAS ========================
        col_g3, col_g4 = st.columns(2)
        
        # ======================== GRÁFICO 6 ========================
        with col_g3:
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.subheader("📊 Órdenes Generadas por Área")
            
            ordenes_por_area = df_filtrado['Area'].value_counts().reset_index()
            ordenes_por_area.columns = ['Área', 'Cantidad']
            ordenes_por_area = ordenes_por_area.sort_values('Cantidad', ascending=False)
            
            fig6, ax6 = plt.subplots(figsize=(10, 5))
            bars6 = ax6.bar(ordenes_por_area['Área'], ordenes_por_area['Cantidad'], color='#7c3aed')
            
            ax6.set_xlabel('Área')
            ax6.set_ylabel('Cantidad')
            ax6.set_title('Órdenes Generadas por Área')
            ax6.set_xticklabels(ordenes_por_area['Área'], rotation=30, ha='right', fontsize=9)
            
            for bar in bars6:
                height = bar.get_height()
                ax6.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                        f'{int(height)}', ha='center', va='bottom', fontsize=11, fontweight='bold', color='black')
            
            plt.tight_layout()
            st.pyplot(fig6)
            
            # Interpretación GRÁFICO 6
            if len(ordenes_por_area) > 0:
                total_ordenes = ordenes_por_area['Cantidad'].sum()
                top_area = ordenes_por_area.iloc[0]['Área']
                top_cantidad = ordenes_por_area.iloc[0]['Cantidad']
                
                texto_interpretacion6 = f'Se generaron <strong>{total_ordenes}</strong> órdenes distribuidas en <strong>{len(ordenes_por_area)}</strong> áreas. El área con mayor generación de órdenes es <span class="stat">"{top_area}"</span> con <strong>{top_cantidad}</strong> órdenes (<span class="stat">{top_cantidad/total_ordenes*100:.1f}%</span> del total).'
                
                if len(ordenes_por_area) > 1:
                    segunda_area = ordenes_por_area.iloc[1]['Área']
                    segunda_cantidad = ordenes_por_area.iloc[1]['Cantidad']
                    texto_interpretacion6 += f' {segunda_area} generó <strong>{segunda_cantidad}</strong> órdenes, representando el <span class="stat">{segunda_cantidad/total_ordenes*100:.1f}%</span> del total.'
                
                st.markdown(generar_interpretacion("Interpretación", texto_interpretacion6), unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICO 7 ========================
        with col_g4:
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.subheader("📊 Estados de Servicios")
            
            estados_counts = df_filtrado['Estado'].value_counts().reset_index()
            estados_counts.columns = ['Estado', 'Cantidad']
            estados_counts = estados_counts.sort_values('Cantidad', ascending=False)
            
            fig7, ax7 = plt.subplots(figsize=(10, 5))
            bars7 = ax7.bar(estados_counts['Estado'], estados_counts['Cantidad'], color='#8b5cf6')
            
            ax7.set_xlabel('Estado')
            ax7.set_ylabel('Cantidad')
            ax7.set_title('Estados de Servicios')
            ax7.set_xticklabels(estados_counts['Estado'], rotation=30, ha='right', fontsize=9)
            
            for bar in bars7:
                height = bar.get_height()
                ax7.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                        f'{int(height)}', ha='center', va='bottom', fontsize=11, fontweight='bold', color='black')
            
            plt.tight_layout()
            st.pyplot(fig7)
            
            # Interpretación GRÁFICO 7
            total_estados_serv = estados_counts['Cantidad'].sum()
            top_estado = estados_counts.iloc[0]['Estado']
            top_estado_cant = estados_counts.iloc[0]['Cantidad']
            
            texto_interpretacion7 = f'El estado más frecuente es <span class="stat">"{top_estado}"</span> con <strong>{top_estado_cant}</strong> órdenes (<span class="stat">{top_estado_cant/total_estados_serv*100:.1f}%</span> del total).'
            
            if len(estados_counts) > 1:
                segundo_estado = estados_counts.iloc[1]['Estado']
                segundo_cantidad = estados_counts.iloc[1]['Cantidad']
                texto_interpretacion7 += f' {segundo_estado} es el segundo estado con <strong>{segundo_cantidad}</strong> órdenes (<span class="stat">{segundo_cantidad/total_estados_serv*100:.1f}%</span> del total).'
            
            texto_interpretacion7 += f' Esto indica que la mayoría de las órdenes se encuentran en estado <span class="stat">"{top_estado}"</span>.'
            
            st.markdown(generar_interpretacion("Interpretación", texto_interpretacion7), unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== GRÁFICO 8 ========================
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.subheader("📊 Ordenamientos Distribuidos por Entidad")
        
        entidad_counts = df_filtrado['Entidad'].value_counts().reset_index()
        entidad_counts.columns = ['Entidad', 'Cantidad']
        entidad_counts = entidad_counts.sort_values('Cantidad', ascending=False)
        
        fig8, ax8 = plt.subplots(figsize=(14, 7))
        bars8 = ax8.bar(entidad_counts['Entidad'], entidad_counts['Cantidad'], color='#6d28d9')
        
        ax8.set_xlabel('Entidad', fontsize=12)
        ax8.set_ylabel('Cantidad', fontsize=12)
        ax8.set_title('Ordenamientos Distribuidos por Entidad', fontsize=14, fontweight='bold')
        ax8.set_xticklabels(entidad_counts['Entidad'], rotation=45, ha='right', fontsize=9)
        
        for bar in bars8:
            height = bar.get_height()
            ax8.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                    f'{int(height)}', ha='center', va='bottom', fontsize=10, fontweight='bold', color='black')
        
        plt.tight_layout()
        st.pyplot(fig8)
        
        # Interpretación GRÁFICO 8
        if len(entidad_counts) > 0:
            total_entidad = entidad_counts['Cantidad'].sum()
            top_entidad = entidad_counts.iloc[0]['Entidad']
            top_entidad_cant = entidad_counts.iloc[0]['Cantidad']
            
            texto_interpretacion8 = f'<strong>{total_entidad}</strong> órdenes están distribuidas entre <strong>{len(entidad_counts)}</strong> entidades. La entidad con mayor volumen es <span class="stat">"{top_entidad}"</span> con <strong>{top_entidad_cant}</strong> órdenes (<span class="stat">{top_entidad_cant/total_entidad*100:.1f}%</span> del total).'
            
            if len(entidad_counts) > 1:
                segunda_entidad = entidad_counts.iloc[1]['Entidad']
                segunda_cantidad = entidad_counts.iloc[1]['Cantidad']
                texto_interpretacion8 += f' {segunda_entidad} es la segunda entidad con <strong>{segunda_cantidad}</strong> órdenes (<span class="stat">{segunda_cantidad/total_entidad*100:.1f}%</span> del total).'
            
            st.markdown(generar_interpretacion("Interpretación", texto_interpretacion8), unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # ======================== EXPORTACIÓN A EXCEL ========================
        st.divider()
        st.markdown("### 📥 Exportar Reporte Completo")
        
        def preparar_datos_exportacion(df_export, df_graf1_data, estado_gestion_data, pendientes_data, 
                                       ordenes_area_data, estados_serv_data, entidad_data, 
                                       df_externas_export, df_externas_proceso):
            datos_detalle = df_export[['Estado', 'Estado_Gestion', 'Solicitado', 'Doc.', 'Paciente', 'Entidad', 'Area', 'Cups', 'Servicio', 'Observación']].copy()
            datos_detalle['Solicitado'] = datos_detalle['Solicitado'].dt.strftime('%Y-%m-%d %H:%M')
            
            resumen_data = []
            
            # Gráfico 1
            total_generadas = df_graf1_data['Generadas'].sum()
            total_gestionadas = df_graf1_data['Gestionadas'].sum()
            pct_gestionadas = (total_gestionadas / total_generadas * 100) if total_generadas > 0 else 0
            resumen_data.append(['Gráfico 1', 'Órdenes Generadas vs Gestionadas', f'Total generadas: {int(total_generadas)}', ''])
            resumen_data.append(['', '', f'Total gestionadas: {int(total_gestionadas)} ({pct_gestionadas:.1f}%)', ''])
            resumen_data.append(['', '', f'Pendientes: {int(total_generadas - total_gestionadas)} ({100-pct_gestionadas:.1f}%)', ''])
            if len(df_graf1_data) > 0:
                resumen_data.append(['', '', f'Día pico: {df_graf1_data.loc[df_graf1_data["Generadas"].idxmax(), "Fecha"]} ({int(df_graf1_data["Generadas"].max())} órdenes)', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 2
            if len(estado_gestion_data) > 0:
                total_estados = estado_gestion_data['Cantidad'].sum()
                for _, row in estado_gestion_data.iterrows():
                    resumen_data.append(['Gráfico 2', 'Gestión de autorizaciones', f'{row["Estado"]}: {row["Cantidad"]} ({row["Cantidad"]/total_estados*100:.1f}%)', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 3
            if len(pendientes_data) > 0:
                total_pend = pendientes_data['Cantidad'].sum()
                for _, row in pendientes_data.iterrows():
                    resumen_data.append(['Gráfico 3', 'Pendientes por Área', f'{row["Área"]}: {row["Cantidad"]} ({row["Cantidad"]/total_pend*100:.1f}%)', ''])
            else:
                resumen_data.append(['Gráfico 3', 'Pendientes por Área', 'No hay datos', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 4 - Solicitudes Externas por Estado
            if df_externas_export is not None and len(df_externas_export) > 0:
                total_ext = df_externas_export['Cantidad'].sum()
                if total_ext > 0:
                    for _, row in df_externas_export.iterrows():
                        resumen_data.append(['Gráfico 4', 'Solicitudes Externas por Estado', f'{row["Estado"]}: {int(row["Cantidad"])} ({row["Cantidad"]/total_ext*100:.1f}%)', ''])
                else:
                    resumen_data.append(['Gráfico 4', 'Solicitudes Externas por Estado', 'No hay datos con cantidad > 0', ''])
            else:
                resumen_data.append(['Gráfico 4', 'Solicitudes Externas por Estado', 'No hay datos', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 5 - Solicitudes Externas por Proceso
            if df_externas_proceso is not None and len(df_externas_proceso) > 0:
                total_proc = df_externas_proceso['Cantidad'].sum()
                if total_proc > 0:
                    for _, row in df_externas_proceso.iterrows():
                        resumen_data.append(['Gráfico 5', 'Solicitudes Externas por Proceso', f'{row["Proceso"]}: {int(row["Cantidad"])} ({row["Cantidad"]/total_proc*100:.1f}%)', ''])
                else:
                    resumen_data.append(['Gráfico 5', 'Solicitudes Externas por Proceso', 'No hay datos', ''])
            else:
                resumen_data.append(['Gráfico 5', 'Solicitudes Externas por Proceso', 'No hay datos', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 6
            if len(ordenes_area_data) > 0:
                total_ord = ordenes_area_data['Cantidad'].sum()
                for _, row in ordenes_area_data.iterrows():
                    resumen_data.append(['Gráfico 6', 'Órdenes por Área', f'{row["Área"]}: {row["Cantidad"]} ({row["Cantidad"]/total_ord*100:.1f}%)', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 7
            if len(estados_serv_data) > 0:
                total_est = estados_serv_data['Cantidad'].sum()
                for _, row in estados_serv_data.iterrows():
                    resumen_data.append(['Gráfico 7', 'Estados de Servicios', f'{row["Estado"]}: {row["Cantidad"]} ({row["Cantidad"]/total_est*100:.1f}%)', ''])
            resumen_data.append(['', '', '', ''])
            
            # Gráfico 8
            if len(entidad_data) > 0:
                total_ent = entidad_data['Cantidad'].sum()
                for _, row in entidad_data.iterrows():
                    resumen_data.append(['Gráfico 8', 'Distribución por Entidad', f'{row["Entidad"]}: {row["Cantidad"]} ({row["Cantidad"]/total_ent*100:.1f}%)', ''])
            
            resumen_df = pd.DataFrame(resumen_data, columns=['Gráfico', 'Categoría', 'Detalle', 'Observación'])
            
            return datos_detalle, resumen_df
        
        if st.button("📥 Exportar Reporte a Excel", use_container_width=True, type="primary"):
            try:
                df_export = df_filtrado.copy()
                
                estado_gestion_data = estado_gestion_counts.copy()
                pendientes_data = pendientes_por_area.copy() if len(df_pendientes) > 0 else pd.DataFrame()
                ordenes_area_data = ordenes_por_area.copy() if len(ordenes_por_area) > 0 else pd.DataFrame()
                estados_serv_data = estados_counts.copy()
                entidad_data = entidad_counts.copy()
                
                # Preparar datos de solicitudes externas para exportación
                df_externas_export = None
                df_externas_proceso = None
                
                if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
                    # Por estado
                    externas_estado = df_externas_filtrado['estado'].value_counts().reset_index()
                    externas_estado.columns = ['Estado', 'Cantidad']
                    df_externas_export = externas_estado
                    
                    # Por proceso
                    if 'proceso' in df_externas_filtrado.columns:
                        externas_proceso = df_externas_filtrado['proceso'].value_counts().reset_index()
                        externas_proceso.columns = ['Proceso', 'Cantidad']
                        df_externas_proceso = externas_proceso
                
                datos_detalle, resumen_graficos = preparar_datos_exportacion(
                    df_export, df_graf1, estado_gestion_data, pendientes_data, 
                    ordenes_area_data, estados_serv_data, entidad_data,
                    df_externas_export, df_externas_proceso
                )
                
                output = BytesIO()
                wb = Workbook()
                
                # Hoja 1: Datos Detallados
                ws1 = wb.active
                ws1.title = "Datos Detallados"
                
                header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for r_idx, row in enumerate(dataframe_to_rows(datos_detalle, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                
                for column in ws1.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_length = min(max_length + 2, 50)
                    ws1.column_dimensions[column_letter].width = adjusted_length
                
                # Hoja 2: Resumen Gráficos
                ws2 = wb.create_sheet("Resumen Gráficos")
                for r_idx, row in enumerate(dataframe_to_rows(resumen_graficos, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                
                for column in ws2.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_length = min(max_length + 2, 50)
                    ws2.column_dimensions[column_letter].width = adjusted_length
                
                # Hoja 3: Solicitudes Externas (datos detallados)
                if df_externas_filtrado is not None and len(df_externas_filtrado) > 0:
                    ws3 = wb.create_sheet("Solicitudes Externas")
                    columnas_ext_export = ['fechaRegistroFormulario', 'ciudad', 'proceso', 'idPaciente', 
                                          'nombrePaciente', 'entidad', 'estado', 'Clasificación']
                    
                    # Crear DataFrame con clasificación
                    df_ext_export = df_externas_filtrado[columnas_ext_export[:-1]].copy() if all(col in df_externas_filtrado.columns for col in columnas_ext_export[:-1]) else pd.DataFrame()
                    if not df_ext_export.empty:
                        df_ext_export['Clasificación'] = df_ext_export['estado'].apply(clasificar_gestion_externa)
                        
                        # Formatear fechas
                        if 'fechaRegistroFormulario' in df_ext_export.columns:
                            df_ext_export['fechaRegistroFormulario'] = pd.to_datetime(df_ext_export['fechaRegistroFormulario']).dt.strftime('%Y-%m-%d')
                        
                        for r_idx, row in enumerate(dataframe_to_rows(df_ext_export, index=False, header=True), 1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                                if r_idx == 1:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.border = thin_border
                        
                        for column in ws3.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_length = min(max_length + 2, 50)
                            ws3.column_dimensions[column_letter].width = adjusted_length
                
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=output,
                    file_name=f"Reporte_Portafolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.success("✅ Reporte generado correctamente. Haz clic en 'Descargar Excel' para guardar el archivo.")
                
            except Exception as e:
                st.error(f"❌ Error al exportar: {e}")
                import traceback
                st.error(traceback.format_exc())
        
        # ======================== INFORMACIÓN DE FILTROS ========================
        st.divider()
        st.caption(f"🔍 Filtros aplicados: {len(estados_seleccionados)} estados, {len(entidades_seleccionadas)} entidades, {len(areas_seleccionadas)} áreas, {len(sedes_seleccionadas)} sedes")
        st.caption(f"📅 Rango de fechas: {fecha_inicio} - {fecha_fin}")
        
else:
    # Mensaje cuando no hay archivo cargado
    st.info("👈 Carga un archivo Excel que contenga las hojas 'Datos', 'Portafolio' y 'Solicitudes Externas' para comenzar a visualizar el dashboard")
    
    with st.expander("📚 Ver Formato Esperado del Archivo", expanded=False):
        st.markdown("""
        ### El archivo Excel debe contener tres hojas:
        
        **Hoja 1: 'Datos'** - Debe contener las siguientes columnas:
        - Tag, Solicitado, Auditado, Sede, Doc., Paciente, Edad, Genero, Diag., Entidad, Grupo Atención, Servicio, Cups, Radicación, Radicado, Autorizado, Autorización, Vence, Entregado, Servicio, Programado, Responsable, Estado, Observación, Prioridad, idOrden, idIndigo
        
        **Hoja 2: 'Portafolio'** - Debe contener las siguientes columnas:
        - CUPS, codIPS, descrCodIPS, codREPS, A, UNIDAD EJECUTORA, Codigo unidad, Sede_Portafolio
        
        **Hoja 3: 'Solicitudes Externas'** - Debe contener las siguientes columnas:
        - fechaRegistroFormulario, ciudad, proceso, idPaciente, nombrePaciente, entidad, servicio, cups, estado, fechaEntregaProceso, motivoCancelacion
        """)

st.divider()
st.caption("💡 Tablero resumen de gestión de Autorizaciones y Programación en Tramita - Datos actualizados en tiempo real")
